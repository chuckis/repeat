"""
Final School Timetabling with OR-Tools CP-SAT
---------------------------------------------
This script unifies Phase 1 (teacher assignment) and Phase 2 (timetable building).
"""
from ortools.sat.python import cp_model
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from subjects_translation import translate_subject, translate_subjects_list

# ------------------------
# DATA
# ------------------------
teachers = ['T1','T2','T3','T4','T5','T6','T7', 'T8', 'T9', 'T10', 'T11',
            'T12','T13','T14', 'T15', 'T17', 'T18', 'T19', 'T20', 'T21', 'T22', 'T23', 'T24']
subjects = ['arithm', 'math', 'algebra', 'geometry', 'ukrmol', 'ukrmollit', 'ukrm', 'ukrlit','english', 'engmol','IT','biology',
            'history', 'ukrhistory','arts','music','crafts', 'craftsboys','sport',
            'physics','geo','pravozn', 'chem', 'prirod', 'ippoter', 'verhova', 'navch',
            'CSL', 'OPK', 'JS', 'event']
classes = [1, 2, 3, 4, 5, 6, 7, 8, 9]
days = ['Mo', 'Tu', 'We', 'Th', 'Fr']
H = 8
lessons = list(range(1, H+1))

# classes hours of subj
Curriculum = dict()

Curriculum[(1, 'event')] = 1
Curriculum[(1, 'ukrmol')] = 6
Curriculum[(1, 'engmol')] = 3
Curriculum[(1, 'arithm')] = 6
Curriculum[(1, 'prirod')] = 3
Curriculum[(1, 'arts')] = 1
Curriculum[(1, 'music')] = 1
# Curriculum[(1, 'crafts')] = 1
Curriculum[(1, 'craftsboys')] = 1
Curriculum[(1, 'sport')] = 2
Curriculum[(1, 'OPK')] = 1
Curriculum[(1, 'JS')] = 1
Curriculum[(1, 'ippoter')] = 1
Curriculum[(1, 'navch')] = 4
# Curriculum[(1, 'CSL')] = 1

Curriculum[(2, 'event')] = 1
Curriculum[(2, 'ukrmol')] = 7
Curriculum[(2, 'engmol')] = 3
Curriculum[(2, 'arithm')] = 6
Curriculum[(2, 'arts')] = 1
Curriculum[(2, 'music')] = 1
# Curriculum[(2, 'crafts')] = 1
Curriculum[(2, 'craftsboys')] = 1
Curriculum[(2, 'sport')] = 2
Curriculum[(2, 'navch')] = 4
Curriculum[(2, 'IT')] = 1
Curriculum[(2, 'prirod')] = 3
Curriculum[(2, 'ippoter')] = 1
Curriculum[(2, 'CSL')] = 1
Curriculum[(2, 'OPK')] = 1
Curriculum[(2, 'JS')] = 1

Curriculum[(3, 'event')] = 1
Curriculum[(3, 'ukrmol')] = 6
Curriculum[(3, 'engmol')] = 3
Curriculum[(3, 'arithm')] = 6
Curriculum[(3, 'prirod')] = 3
Curriculum[(3, 'arts')] = 1
Curriculum[(3, 'music')] = 1
# Curriculum[(3, 'crafts')] = 1
Curriculum[(3, 'craftsboys')] = 1
Curriculum[(3, 'sport')] = 2
Curriculum[(3, 'IT')] = 1
Curriculum[(3, 'ippoter')] = 1
Curriculum[(3, 'CSL')] = 1
Curriculum[(3, 'OPK')] = 1
Curriculum[(3, 'JS')] = 1

Curriculum[(4, 'event')] = 1
Curriculum[(4, 'ukrmol')] = 7
Curriculum[(4, 'engmol')] = 3
Curriculum[(4, 'arithm')] = 6
Curriculum[(4, 'arts')] = 1
Curriculum[(4, 'prirod')] = 3
Curriculum[(4, 'music')] = 1
# Curriculum[(4, 'crafts')] = 1
Curriculum[(4, 'craftsboys')] = 1
Curriculum[(4, 'sport')] = 2
Curriculum[(4, 'IT')] = 1
Curriculum[(4, 'ippoter')] = 1
Curriculum[(4, 'CSL')] = 1
Curriculum[(4, 'OPK')] = 1
Curriculum[(4, 'JS')] = 1

Curriculum[(5, 'event')] = 1  
Curriculum[(5, 'ukrm')] = 4
Curriculum[(5, 'ukrlit')] = 3
Curriculum[(5, 'english')] = 3
Curriculum[(5, 'math')] = 5
Curriculum[(5, 'IT')] = 1
Curriculum[(5, 'geo')] = 2 #TODO: maybe it should be 'prirod'?
Curriculum[(5, 'arts')] = 1
Curriculum[(5, 'history')] = 1
Curriculum[(5, 'music')] = 1
# Curriculum[(5, 'crafts')] = 1
Curriculum[(5, 'craftsboys')] = 1
Curriculum[(5, 'sport')] = 2
Curriculum[(5, 'ippoter')] = 1
Curriculum[(5, 'CSL')] = 1
Curriculum[(5, 'OPK')] = 1
Curriculum[(5, 'JS')] = 1

Curriculum[(6, 'event')] = 1
Curriculum[(6, 'ukrm')] = 4
Curriculum[(6, 'ukrlit')] = 3
Curriculum[(6, 'english')] = 3
Curriculum[(6, 'math')] = 4
Curriculum[(6, 'IT')] = 1
Curriculum[(6, 'biology')] = 2
Curriculum[(6, 'history')] = 1
Curriculum[(6, 'geo')] = 2
Curriculum[(6, 'arts')] = 2
Curriculum[(6, 'music')] = 1
# Curriculum[(6, 'crafts')] = 1
Curriculum[(6, 'craftsboys')] = 1
Curriculum[(6, 'sport')] = 2
Curriculum[(6, 'ippoter')] = 1
Curriculum[(6, 'JS')] = 1
Curriculum[(6, 'OPK')] = 1

Curriculum[(7, 'event')] = 1
Curriculum[(7, 'ukrm')] = 4
Curriculum[(7, 'ukrlit')] = 3
Curriculum[(7, 'english')] = 3
Curriculum[(7, 'math')] = 4
Curriculum[(7, 'IT')] = 1
Curriculum[(7, 'biology')] = 2
Curriculum[(7, 'physics')] = 2
Curriculum[(7, 'history')] = 2 # merged with 8
Curriculum[(7, 'geo')] = 2
Curriculum[(7, 'arts')] = 1
Curriculum[(7, 'music')] = 1
# Curriculum[(7, 'crafts')] = 1
Curriculum[(7, 'craftsboys')] = 1
Curriculum[(7, 'chem')] = 2
Curriculum[(7, 'sport')] = 2
Curriculum[(7, 'JS')] = 1
Curriculum[(7, 'OPK')] = 1
Curriculum[(7, 'ippoter')] = 1

Curriculum[(8, 'event')] = 1
Curriculum[(8, 'ukrm')] = 4
Curriculum[(8, 'ukrlit')] = 3
Curriculum[(8, 'english')] = 3
Curriculum[(8, 'math')] = 5
Curriculum[(8, 'biology')] = 2
Curriculum[(8, 'physics')] = 2
Curriculum[(8, 'history')] = 2
Curriculum[(8, 'chem')] = 2
Curriculum[(8, 'geo')] = 2
Curriculum[(8, 'sport')] = 2
# Curriculum[(8, 'crafts')] = 1
Curriculum[(8, 'craftsboys')] = 1
Curriculum[(8, 'IT')] = 2
Curriculum[(8, 'ippoter')] = 1
Curriculum[(8, 'JS')] = 1
Curriculum[(8, 'OPK')] = 1

Curriculum[(9, 'event')] = 1
Curriculum[(9, 'ukrm')] = 4
Curriculum[(9, 'ukrlit')] = 3
Curriculum[(9, 'english')] = 3
Curriculum[(9, 'math')] = 4
Curriculum[(9, 'biology')] = 2
Curriculum[(9, 'pravozn')] = 1
Curriculum[(9, 'physics')] = 2
Curriculum[(9, 'history')] = 2
Curriculum[(9, 'chem')] = 2
Curriculum[(9, 'geo')] = 1
Curriculum[(9, 'music')] = 1
Curriculum[(9, 'sport')] = 2
Curriculum[(9, 'IT')] = 2
Curriculum[(9, 'ippoter')] = 1
Curriculum[(9, 'OPK')] = 1
# Curriculum[(9, 'crafts')] = 1
Curriculum[(9, 'craftsboys')] = 1


# Approbation (кто какие предметы может вести) - ИСПРАВЛЕНО
Approbation = {
    # Начальные классы - основные учителя
    ('T1','navch'):1,
    ('T1','prirod'):1,
    ('T1','ukrmol'):1,
    ('T1','arithm'):1,
    ('T1','crafts'):1,  # Добавлено для базовых предметов
    
    ('T2', 'navch'):1,
    ('T2','prirod'):1,
    ('T2','ukrmol'):1,
    ('T2','arithm'):1,
    ('T2','crafts'):1,  # Добавлено для базовых предметов
    
    ('T3', 'ukrm'):1,
    ('T3', 'ukrlit'):1,
    ('T3','ukrmol'):1,
    ('T3','arithm'):1,
    ('T3','prirod'):1,
    ('T3','navch'):1,   # Добавлено для базовых предметов
    ('T3', 'pravozn'):1,
    
    # T4 - музыка + может вести начальные классы
    ('T4', 'music'):1,
    ('T4','arithm'):1,
    ('T4','ukrmol'):1,
    ('T4','prirod'):1,
    ('T4','navch'):1,   # Добавлено для базовых предметов
    
    # Предметники
    ('T5', 'CSL'):1,
    ('T6', 'math'):1,
    ('T6', 'algebra'):1,
    ('T6', 'geometry'):1,
    ('T7', 'arts'):1,
    # ('T7', 'JS'):1,
    ('T8', 'OPK'):1,
    ('T8', 'event'):1,
    ('T9', 'sport'):1,
    ('T10', 'IT'):1,
    ('T11', 'verhova'):1,
    ('T12', 'biology'):1,
    ('T13', 'history'):1,
    ('T24', 'english'):1,
    ('T14', 'engmol'):1,
    ('T15', 'chem'):1,
    # ('T16', 'crafts'):1,
    ('T17', 'craftsboys'):1,
    ('T18', 'JS'):1,
    ('T19', 'physics'):1,
    ('T20', 'ukrm'):1,
    ('T21', 'ippoter'):1,
    ('T22', 'geo'):1,
    ('T23', 'ukrlit'):1,
}


# ------------------------
# PHASE 1: Teacher assignment (fixed for 1-4 classes, but open for others)
# ------------------------
model1 = cp_model.CpModel()

Teaches = {}
for t in teachers:
    for c in classes:
        for s in subjects:
            Teaches[t, c, s] = model1.NewBoolVar(f"Teach[{t},{c},{s}]")

Lessons = {t: model1.NewIntVar(0, 40, f"Lessons[{t}]") for t in teachers}

# ИСПРАВЛЕНО: базовые предметы для 1-4 классов
base_subjects = {"ukrmol", "arithm", "navch", "prirod"} # !!!!!
pairings = list(zip(teachers[:4], classes[:4]))  # (T1,1), (T2,2), (T3,3), (T4,4)

# Закрепляем базовых учителей за своими классами
for t, c in pairings:
    for s in base_subjects:
        if Curriculum.get((c, s), 0) > 0:
            # Проверяем, может ли учитель вести этот предмет
            if Approbation.get((t, s), 0) == 1:
                # закрепляем учителя за этим предметом в "своём" классе
                model1.Add(Teaches[t, c, s] == 1)
                # запрещаем другим учителям этот предмет в этом классе
                # TODO: think about navch in 4th class
                for other_t in teachers:
                    if other_t != t:
                        model1.Add(Teaches[other_t, c, s] == 0)

# остальное распределяем по approbation
for t in teachers:
    total = []
    for c in classes:
        for s in subjects:
            hrs = Curriculum.get((c, s), 0)
            if hrs > 0:
                # event всегда ведёт T8
                if s == 'event':
                    if t == 'T8':
                        total.append(hrs * Teaches[t, c, s])
                        model1.Add(Teaches[t, c, s] == 1)
                    else:
                        model1.Add(Teaches[t, c, s] == 0)
                # остальные предметы по approbation
                elif Approbation.get((t, s), 0) == 1:
                    total.append(hrs * Teaches[t, c, s])
                else:
                    model1.Add(Teaches[t, c, s] == 0)
            else:
                model1.Add(Teaches[t, c, s] == 0)

    if total:
        model1.Add(Lessons[t] == sum(total))
    else:
        model1.Add(Lessons[t] == 0)
    model1.Add(Lessons[t] <= 40)

# каждая пара (class, subject) должна иметь ровно одного учителя
for c in classes:
    for s in subjects:
        hrs = Curriculum.get((c, s), 0)
        if hrs > 0:
            if s == 'event':
                # event всегда ведёт T8
                model1.Add(Teaches['T8', c, s] == 1)
            elif c in [1, 2, 3, 4] and s in base_subjects:
                # базовые предметы уже зафиксированы выше
                pass
            else:
                possible_teachers = [t for t in teachers if Approbation.get((t, s), 0) == 1]
                if not possible_teachers:
                    print(f"❌ ВНИМАНИЕ: Нет учителя для предмета {s} в классе {c}")
                    # Можно добавить исключение или пропустить
                    continue
                model1.Add(sum(Teaches[t, c, s] for t in possible_teachers) == 1)

# решаем Phase 1
solver1 = cp_model.CpSolver()
solver1.parameters.max_time_in_seconds = 20
status1 = solver1.Solve(model1)

teacher_of = {}
if status1 in (cp_model.OPTIMAL, cp_model.FEASIBLE):
    for c in classes:
        for s in subjects:
            for t in teachers:
                if solver1.Value(Teaches[t, c, s]):
                    teacher_of[(c, s)] = t
    print("✅ Phase 1 solved. Teacher assignment ready.")
else:
    print(f"❌ Phase 1 has no feasible solution. Status: {solver1.StatusName(status1)}")
    
    # Попытка диагностики
    print("\n🔍 Диагностика проблем:")
    for c in classes:
        for s in subjects:
            if Curriculum.get((c, s), 0) > 0:
                possible_teachers = [t for t in teachers if Approbation.get((t, s), 0) == 1]
                if not possible_teachers:
                    print(f"  ❌ Класс {c}, предмет {s}: нет подходящих учителей")

# ------------------------
# TEACHER WORKLOAD SUMMARY (после Phase 1)
# ------------------------
from collections import defaultdict

def summarize_teacher_load(teachers, classes, subjects, Curriculum, Lessons, teacher_of, solver1):
    # 1) Часы по модели (Lessons[t])
    hours_by_model = {t: solver1.Value(Lessons[t]) for t in teachers}

    # 2) Пересчёт по назначению teacher_of и Curriculum
    hours_by_calc = defaultdict(int)
    for c in classes:
        for s in subjects:
            hrs = Curriculum.get((c, s), 0)
            if hrs > 0:
                t = teacher_of.get((c, s))
                if t is None:
                    print(f"⚠️ Нет назначенного учителя для ({c}, '{s}') — проверь Phase 1.")
                    continue
                hours_by_calc[t] += hrs

    # Сводная печать (сортировка по teachers)
    print("\n=== Teacher workload (hours) ===")
    print(f"{'Teacher':<8} {'Model':>6} {'Calc':>6}  Note")
    for t in teachers:
        m = hours_by_model.get(t, 0)
        c = hours_by_calc.get(t, 0)
        mark = "" if m == c else "  ⛔ mismatch"
        print(f"{t:<8} {m:>6} {c:>6}  {mark}")

    # Итого по школе (проверка сумм)
    total_model = sum(hours_by_model.values())
    total_calc  = sum(hours_by_calc.values())
    print(f"\nTOTAL  Model={total_model}  Calc={total_calc} {'OK' if total_model==total_calc else '⛔'}")

# вызов сразу после успешного решения Phase 1:
if status1 in (cp_model.OPTIMAL, cp_model.FEASIBLE):
    summarize_teacher_load(teachers, classes, subjects, Curriculum, Lessons, teacher_of, solver1)

# ------------------------
# PHASE 2: Timetable (без кабинетов)
# ------------------------
model2 = cp_model.CpModel()

Timetable = {}
for c in classes:
    for s in subjects:
        for d in days:
            for h in lessons:
                Timetable[c, s, d, h] = model2.NewBoolVar(f"TT[{c},{s},{d},{h}]")

# Curriculum adherence (каждый предмет = заданное количество часов)
for c in classes:
    for s in subjects:
        hrs = Curriculum.get((c, s), 0)
        if hrs > 0:
            model2.Add(sum(Timetable[c, s, d, h] for d in days for h in lessons) == hrs)

# Ограничение: уроки подряд (без дырок)
for c in classes:
    # граница по классам
    if c <= 4:  
        last_lesson = 7
    else:
        last_lesson = 8
    for d in days:
        for h in range(1, last_lesson):
            has_lesson_h   = sum(Timetable[c, s, d, h]   for s in subjects)
            has_lesson_h1  = sum(Timetable[c, s, d, h+1] for s in subjects)
            model2.Add(has_lesson_h >= has_lesson_h1)

# Ограничение: в понедельник на первом уроке все классы имеют event
# event_subject = 'event'
for c in classes:
    # Обязательный урок в понедельник, 1-й урок
    model2.Add(Timetable[c, 'event', 'Mo', 1] == 1)
    # Другие слоты в понедельник и в другие дни запрещаем для event
    for d in days:
        for h in lessons:
            if (d, h) != ('Mo', 1):
                model2.Add(Timetable[c, 'event', d, h] == 0)


# Ограничение: OPK для классов 1-5 только в любые 2 дня, кроме вторника
# for c in range(1, 6):
#     opk_day_bools = []
#     for d in days:
#         if d == 'Tu':
#             for h in lessons:
#                 model2.Add(Timetable[c, 'OPK', d, h] == 0)
#         else:
#             opk_day = model2.NewBoolVar(f"opk_day_{c}_{d}")
#             opk_day_bools.append(opk_day)
#             slot_sum = sum(Timetable[c, 'OPK', d, h] for h in lessons)
#             model2.Add(slot_sum >= 1).OnlyEnforceIf(opk_day)
#             model2.Add(slot_sum == 0).OnlyEnforceIf(opk_day.Not())
#     model2.Add(sum(opk_day_bools) == 2)

# Ограничение: у 6-9 классов OPK вместе (один слот для всех)
# slots = [(d, h) for d in days for h in lessons]
# opk_slot_bools = []
# for d, h in slots:
#     bool_var = model2.NewBoolVar(f"opk_slot_{d}_{h}")
#     opk_slot_bools.append(bool_var)
#     for c in range(6, 10):
#         model2.Add(Timetable[c, 'OPK', d, h] == 1).OnlyEnforceIf(bool_var)
#         for d2, h2 in slots:
#             if (d2, h2) != (d, h):
#                 model2.Add(Timetable[c, 'OPK', d2, h2] == 0).OnlyEnforceIf(bool_var)
# # мягкий вариант: разрешаем несколько дней, но минимизируем
# # (если хочешь жёстко один слот, оставь sum==1)
# model2.Add(sum(opk_slot_bools) >= 1)

# Ограничение: crafts только на 2, 3, 6 или 7 уроке
# allowed_crafts_lessons = [2, 3, 6, 7]
# for c in classes:
#     for d in days:
#         for h in lessons:
#             if h not in allowed_crafts_lessons:
#                 model2.Add(Timetable[c, 'crafts', d, h] == 0)

# Ограничение: sport только во вторник, четверг и пятницу
# allowed_sport_days = ['Tu', 'Th', 'Fr']
# for c in classes:
#     for d in days:
#         if d not in allowed_sport_days:
#             for h in lessons:
#                 model2.Add(Timetable[c, 'sport', d, h] == 0)

# Ограничение: biology только в понедельник и пятницу
# allowed_bio_days = ['Mo', 'Fr']
# for c in classes:
#     for d in days:
#         if d not in allowed_bio_days:
#             for h in lessons:
#                 model2.Add(Timetable[c, 'biology', d, h] == 0)

# Ограничение: ippoter в 5-9 классах только в среду
# for c in range(6, 10):
#     for d in days:
#         if d != 'We':
#             for h in lessons:
#                 model2.Add(Timetable[c, 'ippoter', d, h] == 0)

# Ограничение: один учитель ≤ 1 урок в слот (кроме совместных OPK и 7+8 вместе)
for t in teachers:
    for d in days:
        for h in lessons:
            normal_lessons = []
            for c in classes:
                for s in subjects:
                    if teacher_of.get((c, s)) == t:
                        # исключаем совместные OPK (6-9) и пару 7+8
                        if not (s == 'OPK' and 6 <= c <= 9) and not (c == 7 or c == 8):
                            normal_lessons.append(Timetable[c, s, d, h])

#             # теперь добавляем пересечение для 7 и 8 отдельно
#             # все остальные классы должны быть ≤1
            model2.Add(sum(normal_lessons) <= 1)

# Objective: минимизируем расхождения с Curriculum + минимизация числа дней для OPK (6–9)
penalties = []
for c in classes:
    for s in subjects:
        hrs = Curriculum.get((c, s), 0)
        if hrs > 0:
            actual = sum(Timetable[c, s, d, h] for d in days for h in lessons)
            diff = model2.NewIntVar(-hrs, hrs, f"diff_{c}_{s}")
            abs_diff = model2.NewIntVar(0, hrs, f"abs_diff_{c}_{s}")
            model2.Add(diff == actual - hrs)
            model2.AddAbsEquality(abs_diff, diff)
            penalties.append(abs_diff)

# 2) Мягкое ограничение для OPK (1-5 классы)
# for c in range(1, 6):
#     opk_day_bools = []
#     for d in days:
#         if d == 'Tu':  # вторник запрещён
#             for h in lessons:
#                 model2.Add(Timetable[c, 'OPK', d, h] == 0)
#             continue

#         opk_day = model2.NewBoolVar(f"opk_day_{c}_{d}")
#         opk_day_bools.append(opk_day)
#         slot_sum = sum(Timetable[c, 'OPK', d, h] for h in lessons)
#         model2.Add(slot_sum >= 1).OnlyEnforceIf(opk_day)
#         model2.Add(slot_sum == 0).OnlyEnforceIf(opk_day.Not())

#     # ограничение: не более 2 дней с ОПК
#     model2.Add(sum(opk_day_bools) <= 2)

#     # добавляем в цель, чтобы минимизировать количество дней с ОПК
#     penalties.append(sum(opk_day_bools))

model2.Minimize(sum(penalties)) # + sum(opk_day_bools))

# Решение Phase 2
solver2 = cp_model.CpSolver()
solver2.parameters.max_time_in_seconds = 180
status2 = solver2.Solve(model2) # TODO: change to SolveWithSolutionCallback for intermediate results

# ------------------------
# VISUALIZATION
# ------------------------
import matplotlib.pyplot as plt

def show_timetable_table(Timetable, classes, subjects, days, lessons, solver):
    fig, ax = plt.subplots(figsize=(len(classes)*1.5, len(days)*len(lessons)*0.4))

    # соответствие английских кодов дней к укр. сокращениям
    day_labels = {
        "Mo": "Пн",
        "Tu": "Вт",
        "We": "Ср",
        "Th": "Чт",
        "Fr": "Пт"
    }

    # всего строк = дни * уроки
    n_rows = len(days) * len(lessons)
    n_cols = len(classes) + 2  # +2: колонка "День", колонка "№ урока"

    # пустая таблица
    cell_text = [["" for _ in range(n_cols)] for _ in range(n_rows)]

    # заголовки
    col_labels = ["День", "№"] + [str(c) for c in classes]

    # заполняем таблицу
    row = 0
    for d in days:
        for hi, h in enumerate(lessons):
            # День пишем только в первой строке блока
            cell_text[row][0] = day_labels.get(d, d) if hi == 0 else ""
            cell_text[row][1] = str(h)  # номер урока
            for ci, c in enumerate(classes):
                subj = None
                for s in subjects:
                    if solver.Value(Timetable[c, s, d, h]):
                        subj = translate_subject(s)
                cell_text[row][ci+2] = subj if subj else ""
            row += 1

    # строим таблицу matplotlib
    the_table = ax.table(cellText=cell_text, colLabels=col_labels, loc='center',
                         cellLoc='center', edges='closed')

    the_table.auto_set_font_size(False)
    the_table.set_fontsize(7)
    the_table.scale(1.2, 1.2)

    # жирные линии-разделители между днями
    for di in range(1, len(days)):
        row_index = di * len(lessons)
        for col in range(n_cols):
            the_table[(row_index, col)].visible_edges = "T"
            the_table[(row_index, col)].set_linewidth(2.0)

    ax.axis('off')
    plt.tight_layout()
    plt.show()

def export_timetable_to_excel(filename, Timetable, classes, subjects, days, lessons, solver):
    # соответствие кодов дней к укр. сокращениям
    day_labels = {
        "Mo": "Пн",
        "Tu": "Вт",
        "We": "Ср",
        "Th": "Чт",
        "Fr": "Пт"
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "Розклад"

    # заголовки
    headers = ["День", "№"] + [str(c) for c in classes]
    ws.append(headers)

    # заполняем таблицу
    for d in days:
        for hi, h in enumerate(lessons):
            row = []
            row.append(day_labels.get(d, d) if hi == 0 else "")
            row.append(h)
            for c in classes:
                subj = None
                for s in subjects:
                    if solver.Value(Timetable[c, s, d, h]):
                        subj = translate_subject(s)
                row.append(subj if subj else "")
            ws.append(row)

    # Стилизация таблицы
    thin = Side(border_style="thin", color="000000")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    # Заголовки жирным
    for cell in ws[1]:
        cell.font = Font(bold=True)

    wb.save(filename)
    print(f"✅ Розклад збережено у файл {filename}")


if status2 in (cp_model.OPTIMAL, cp_model.FEASIBLE):
    print("✅ Feasible timetable found")
   # Визуализация как таблица в matplotlib
    show_timetable_table(Timetable, classes, subjects, days, lessons, solver2)

# Экспорт в Excel
    # export_timetable_to_excel("timetable.xlsx", Timetable, classes, subjects, days, lessons, solver2)

else:
    print("❌ No feasible timetable")